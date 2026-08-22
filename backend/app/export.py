from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def build_jobs_workbook(jobs) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jobs"
    headers = [
        "Score",
        "Title",
        "Company",
        "ATS Platform",
        "Status",
        "Apply URL",
        "Missing Skills",
        "Weak Requirements",
        "Date Fetched",
        "Date Applied",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for job in sorted(jobs, key=lambda item: item.score, reverse=True):
        sheet.append(
            [
                job.score,
                job.title,
                job.company,
                job.ats_platform,
                job.status.value,
                job.apply_url,
                ", ".join(job.missing_skills or []),
                "; ".join(job.weak_requirements or []),
                job.date_fetched.replace(tzinfo=None) if job.date_fetched else None,
                job.date_applied.replace(tzinfo=None) if job.date_applied else None,
            ]
        )
        link = sheet.cell(sheet.max_row, 6)
        if job.apply_url:
            link.hyperlink = job.apply_url
            link.style = "Hyperlink"
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[column[0].column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
