"""Parses an uploaded .xlsx/.csv of jobs for an upload-mode batch.

The expected column layout is exactly what `export.py`'s "Download Excel" produces
(Score, Title, Company, ATS Platform, Status, Apply URL, Missing Skills, Weak
Requirements, Date Fetched, Date Applied) — the intended workflow is export, review
or edit offline, then re-upload to auto-apply unattended to whatever still clears the
threshold. Three extra columns (Requirement Coverage, Skill Coverage, Global
Similarity) are accepted if present but aren't part of the export format, so they
default to 0 when absent.

Status and Date Applied columns are tolerated (so re-uploading a file this app
itself exported never errors on an unrecognized header) but their values are
deliberately never read: every upload batch re-applies to every row it's given,
regardless of what either said last time a previous run touched it — an explicit
"apply all of these, no matter what" per row, not a resume-where-you-left-off. See
upsert_jobs_from_rows.
"""

import csv
import hashlib
import io
from datetime import datetime, timezone
from urllib.parse import urlparse

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .extractor import ATS_DOMAINS
from .models import Job, JobStatus

REQUIRED_HEADERS = {"title", "company", "apply url"}
HEADER_FIELDS = {
    "score": "score",
    "title": "title",
    "company": "company",
    "ats platform": "ats_platform",
    # No "status" entry: a Status column is still tolerated in the file (an
    # unrecognized header is silently ignored, never an error — see parse_sheet),
    # but its value is never applied to the Job row. See this module's docstring
    # and upsert_jobs_from_rows for why.
    "apply url": "apply_url",
    "missing skills": "missing_skills",
    "weak requirements": "weak_requirements",
    "date fetched": "date_fetched",
    # No "date applied" entry either, for the same reason as Status above: it's
    # always force-reset to None on upload, in lockstep with status always
    # resetting to discovered (see upsert_jobs_from_rows).
    "requirement coverage": "requirement_coverage",
    "skill coverage": "skill_coverage",
    "global similarity": "global_similarity",
}


class SheetImportError(RuntimeError):
    pass


def _normalize_header(value) -> str:
    return str(value or "").strip().lower()


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _split_list(value, separator: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(separator) if item.strip()]


def _parse_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _derive_ats(apply_url: str) -> str:
    domain = urlparse(apply_url).netloc.lower()
    return next((name for marker, name in ATS_DOMAINS.items() if marker in domain), "career-page")


def _rows_from_csv(contents: bytes) -> list[list]:
    text = contents.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(cell.strip() for cell in row)]


def _rows_from_xlsx(contents: bytes) -> list[list]:
    workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None for cell in row)
    ]


def parse_sheet(filename: str, contents: bytes) -> list[dict]:
    """Returns a list of field-name -> raw-cell-value dicts, one per row with an apply
    link. Raises SheetImportError for a bad file or a missing required column."""
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        rows = _rows_from_csv(contents)
    elif lower.endswith(".xlsx"):
        rows = _rows_from_xlsx(contents)
    else:
        raise SheetImportError("Please upload a .xlsx or .csv file.")
    if not rows:
        raise SheetImportError("The file is empty.")

    headers, *data_rows = rows
    normalized = [_normalize_header(h) for h in headers]
    missing = REQUIRED_HEADERS - set(normalized)
    if missing:
        raise SheetImportError(f"Missing required column(s): {', '.join(sorted(missing))}")
    fields = [HEADER_FIELDS.get(h) for h in normalized]

    parsed = []
    for raw_row in data_rows:
        record = {field: value for field, value in zip(fields, raw_row) if field}
        if str(record.get("apply_url", "")).strip():
            parsed.append(record)
    return parsed


def import_external_id(apply_url: str) -> str:
    """A stable id for an uploaded row, derived from the job's own apply_url
    rather than the batch it happened to be imported through. `f"import:
    {batch_id}:{index}"` (the original scheme) broke in a real, confirmed way:
    a batch's own primary key can be reused by SQLite once old batch rows are
    gone (no true AUTOINCREMENT here — see migrations.py), so two *separate*
    uploads of the same file could compute the exact same external_id and
    collide on the (workspace_id, external_id) unique constraint — a 500, not
    a graceful outcome. Deriving it from the URL instead means uploading the
    same job twice, in the same workspace, always resolves to the same row —
    which is also exactly what this module's own docstring already describes
    as the intended workflow (export, edit, re-upload) — see
    routes/batches.py::create_upload_batch for the update-in-place handling
    that makes a genuine re-upload work instead of erroring."""
    return f"import:{hashlib.sha256(apply_url.encode('utf-8')).hexdigest()[:24]}"


def upsert_jobs_from_rows(db: Session, rows: list[dict], workspace_id: int, batch_id: int) -> None:
    """Inserts or updates one Job per row, keyed by (workspace_id,
    import_external_id(apply_url)) — a genuine re-upload of the same job (see
    import_external_id's docstring for why that's now detectable at all)
    updates the existing row in place instead of colliding on the unique
    constraint, matching this module's own documented "export, edit,
    re-upload" workflow. Re-entering an existing job into a fresh batch's
    scope also resets auto_apply_state/review_reason to let *that* batch's own
    run decide again, rather than carrying over a decision made by whatever
    batch touched the row last.

    status is always force-reset to discovered here (date_applied along with
    it, matching the same invariant routes/jobs.py::patch_job enforces on a
    manual change) rather than read from the row's own Status/Date Applied
    cells — deliberately, so every upload batch re-applies to every row it's
    given, even one already marked applied (auto *or manual*) the last time it
    was touched. Uploading a specific URL is itself taken as the explicit,
    per-run instruction to (re)attempt it, superseding whatever an earlier
    manual "Mark applied"/"Skip" said about that same row. That manual mark
    still fully matters for a job that only ever shows up via a search-sourced
    batch (see scheduler._run_batch's gate) — it just doesn't carry across a
    later re-upload of the same URL."""
    for row in rows:
        apply_url = str(row.get("apply_url", "")).strip()
        external_id = import_external_id(apply_url)
        job = db.scalar(
            select(Job).where(Job.workspace_id == workspace_id, Job.external_id == external_id)
        )
        if job is None:
            job = Job(workspace_id=workspace_id, external_id=external_id)
            db.add(job)
        job.title = str(row.get("title") or "").strip() or "Untitled role"
        job.company = str(row.get("company") or "").strip() or "Unknown company"
        job.ats_platform = str(row.get("ats_platform") or "").strip().lower() or _derive_ats(
            apply_url
        )
        job.apply_url = apply_url
        job.score = _to_float(row.get("score"))
        job.requirement_coverage = _to_float(row.get("requirement_coverage"))
        job.skill_coverage = _to_float(row.get("skill_coverage"))
        job.global_similarity = _to_float(row.get("global_similarity"))
        job.missing_skills = _split_list(row.get("missing_skills"), ",")
        job.weak_requirements = _split_list(row.get("weak_requirements"), ";")
        job.status = JobStatus.discovered
        job.date_fetched = _parse_datetime(row.get("date_fetched")) or datetime.now(timezone.utc)
        job.date_applied = None
        job.source_batch_id = batch_id
        job.auto_apply_state = None
        job.review_reason = None
